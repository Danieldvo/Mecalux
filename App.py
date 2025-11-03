import pandas as pd
import streamlit as st
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =====================
# 📁 CARGA DE DATOS
# =====================
ruta_carpeta = Path(__file__).parent
nombre_archivo = "df_materiales.xlsx"
hoja_materiales = "BD_Materiales"
ruta_completa = ruta_carpeta / nombre_archivo

df_materiales = pd.read_excel(ruta_completa, sheet_name=hoja_materiales)

# Convertimos la columna PrecioUnit a numérico (por si viene con símbolo €)
df_materiales["PrecioUnit"] = (
    df_materiales["PrecioUnit"]
    .astype(str)
    .str.replace("€", "")
    .str.replace(",", ".")
    .astype(float)
)

# =====================
# 💵 CARGA DE TARIFAS
# =====================
hoja_tarifas = "BD_Tarifas"
df_tarifas = pd.read_excel(ruta_completa, sheet_name=hoja_tarifas)

# Limpiamos posibles espacios o formatos
df_tarifas.columns = df_tarifas.columns.str.strip()
df_tarifas["Técnico"] = df_tarifas["Técnico"].str.strip()
df_tarifas["Tipo"] = df_tarifas["Tipo"].str.strip()

# Convertimos la columna TARIFA a numérico (por si viene con símbolo €)
df_tarifas["TARIFA"] = (
    df_tarifas["TARIFA"]
    .astype(str)
    .str.replace("€", "")
    .str.replace(",", ".")
    .astype(float)
)

# =====================
# 🧩 FUNCIONES DE FILTRO
# =====================
def filtrar_por_red(df, red):
    if red == "Profinet":
        return df[df["Red"].isin(["Ambos", "Profinet"])]
    elif red == "Profibus":
        return df[df["Red"].isin(["Ambos", "Profibus"])]
    return df.copy()

def filtrar_por_funcion(df, funcion):
    if funcion == "Vision&Control":
        return df[df["VisControl"] == 1]
    elif funcion == "Vision":
        return df[df["Vis"] == 1]
    return df.copy()

def filtrar_por_iluminacion(df, iluminacion):
    df_filtrado = df.copy()

    if iluminacion.upper() == "NO":
        df_filtrado = df_filtrado[
            df_filtrado["Descripcion"].str.upper().str.strip() != "FOCO LED 10W 220V 6500K 1150LM"
        ]

    # Si es "SI" o cualquier otro valor, no se toca
    return df_filtrado

def filtrar_por_clima(df, clima):
    if clima == "Ambiente":
        return df[df["Clima"].isin(["Ambos", "Ambiente"])]
    elif clima == "Frio":
        return df[df["Clima"].isin(["Ambos", "Frio"])]
    return df.copy()

def filtrar_por_pc_nuevo(df, pc_nuevo):
    if pc_nuevo == "NO":
        return df[df["Descripcion"].str.upper() != "PC INDUSTRIAL CPC3G"]
    return df.copy()

def tratar_por_distancia(df, distancia, red):
    df_filtrado = df.copy()

    # Actualizar cantidad de la manguera
    mask_manguera = df_filtrado["Descripcion"].str.upper() == "MANGUERA  3G1,0 NUM.-ML"
    df_filtrado.loc[mask_manguera, "Cantidad"] = distancia

    # Eliminar SWITCH si no aplica (solo se mantiene si Profibus y distancia > 150)
    if not (red == "Profibus" and distancia >= 150):
        df_filtrado = df_filtrado[df_filtrado["Descripcion"].str.upper() != "SWITCH ETHERNET GESTIONABLE"]

    return df_filtrado

def tratar_por_num_maquinas(df, num_maq):
    df_filtrado = df.copy()
    df_filtrado["Cantidad"] = pd.to_numeric(df_filtrado["Cantidad"], errors="coerce")
    df_filtrado.loc[df_filtrado["Unico"] == 0, "Cantidad"] *= int(num_maq)
    return df_filtrado

# =====================
# 🎛️ INTERFAZ STREAMLIT
# =====================
st.set_page_config(page_title="📋 Oferta Cámaras Webs", layout="centered")
st.title("📋 Oferta Cámaras Webs")

with st.sidebar:
    st.header("⚙️ Parámetros de configuración")

    red = st.selectbox("Red", ["Profinet", "Profibus"])
    funcion = st.selectbox("Función", ["Vision", "Vision&Control"])
    iluminacion = st.selectbox("Iluminación", ["SI", "NO"])
    clima = st.selectbox("Clima", ["Ambiente", "Frio"])
    pc_nuevo = st.selectbox("PC Nuevo", ["SI", "NO"])
    distancia = st.number_input("Distancia Ft->Arm (m)", min_value=0, max_value=1000, value=100, step=10)
    num_maq = st.number_input("Número de máquinas", min_value=1, max_value=50, value=1, step=1)



    # =====================
    # 👷 BLOQUE DE MANO DE OBRA (SEPARADO POR TÉCNICO)
    # =====================
    st.header("👷 Mano de obra")
    
    # --- Técnico Electromecánico ---
    st.subheader("🧰 Técnico Electromecánico")

    km_electro = st.number_input(
    "Kilómetros desde la delegación a la instalación (ida + vuelta)",
    min_value=0,
    max_value=9000,
    value=0,
    step=10,
    key="km_electro"
    )

    h_electro = st.number_input(
    "Horas de desplazamiento (ida + vuelta)",
    min_value=0,
    max_value=24,
    value=0,
    step=1,
    key="h_electro"
    )

    frio_electro = st.selectbox(
    "Trabajo en frío (<4ºC)",
    ["NO", "SI"],
    key="frio_electro"
    )

    estancia_electro = st.selectbox(
    "Estancia del operario",
    ["VIAJA", "PERNOCTA"],
    key="estancia_electro"
    )

    # --- NUEVAS VARIABLES: trabajo en festivo y nocturno (Electromecánico) ---
    st.markdown("#### 🌙 Horas especiales (Electromecánico)")

    horas_festivo_electro = st.number_input(
    "Horas en festivo",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="horas_festivo_electro"
    )

    horas_nocturno_electro = st.number_input(
    "Horas en nocturno",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="horas_nocturno_electro"
    )
    # --- NUEVAS VARIABLES: Tipo de viaje y vehículo de alquiler (Electromecánico) ---
    st.markdown("#### ✈️ Desplazamientos (Electromecánico)")

    tipo_viaje_electro = st.selectbox(
    "Tipo de viaje",
    ["COCHE", "AVION"],
    key="tipo_viaje_electro"
    )

    dias_alquiler_electro = st.number_input(
    "Vehículo de alquiler (días)",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="dias_alquiler_electro"
    )

    # --- Técnico de Control ---
    st.subheader("🧠 Técnico de Control")
    km_ctrl = st.number_input(
    "Kilómetros desde la delegación a la instalación (ida + vuelta)",
    min_value=0,
    max_value=9000,
    value=0,
    step=10,
    key="km_ctrl"
    )

    h_ctrl = st.number_input(
    "Horas de desplazamiento (ida + vuelta)",
    min_value=0,
    max_value=24,
    value=0,
    step=1,
    key="h_ctrl"
    )

    frio_ctrl = st.selectbox(
    "Trabajo en frío (<4ºC)",
    ["NO", "SI"],
    key="frio_ctrl"
    )
    
    estancia_ctrl = st.selectbox(
    "Estancia del operario",
    ["VIAJA", "PERNOCTA"],
    key="estancia_ctrl"
    )
    # --- NUEVAS VARIABLES: trabajo en festivo y nocturno (Control) ---
    st.markdown("#### 🌙 Horas especiales (Control)")

    horas_festivo_ctrl = st.number_input(
    "Horas en festivo",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="horas_festivo_ctrl"
    )

    horas_nocturno_ctrl = st.number_input(
    "Horas en nocturno",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="horas_nocturno_ctrl"
    )
    # --- NUEVAS VARIABLES: Tipo de viaje y vehículo de alquiler (Control) ---
    st.markdown("#### ✈️ Desplazamientos (Control)")

    tipo_viaje_ctrl = st.selectbox(
    "Tipo de viaje",
    ["COCHE", "AVION"],
    key="tipo_viaje_ctrl"
    )

    dias_alquiler_ctrl = st.number_input(
    "Vehículo de alquiler (días)",
    min_value=0.0,
    step=1.0,
    value=0.0,
    key="dias_alquiler_ctrl"
    )

# =====================
# 🧮 APLICAR FILTROS
# =====================
df_filtrado = filtrar_por_red(df_materiales, red)
df_filtrado = filtrar_por_funcion(df_filtrado, funcion)
df_filtrado = filtrar_por_iluminacion(df_filtrado, iluminacion)
df_filtrado = filtrar_por_clima(df_filtrado, clima)
df_filtrado = filtrar_por_pc_nuevo(df_filtrado, pc_nuevo)
df_filtrado = tratar_por_distancia(df_filtrado, distancia, red)
df_filtrado = tratar_por_num_maquinas(df_filtrado, num_maq)

df_interno = df_filtrado[df_filtrado["Categoria"] == "Interno"]
df_externo = df_filtrado[df_filtrado["Categoria"] == "Externo"]

# =====================
# ➕ AÑADIR MATERIALES ADICIONALES
# =====================
st.markdown("---")
st.subheader("➕ Añadir materiales adicionales")

# Crear contenedor expandible
with st.expander("Agregar material adicional"):
    col1, col2, col3, col4,col5 = st.columns([2,3, 2, 2, 2])
    with col1:
         referencia_extra = st.text_input("Referencia del material", "")
    with col2:
        descripcion_extra = st.text_input("Descripción del material", "")
    with col3:
        cantidad_extra = st.number_input("Cantidad", min_value=0.0, step=1.0, value=0.0)
    with col4:
        precio_unit_extra = st.number_input("Precio unitario (€)", min_value=0.0, step=1.0, value=0.0)
    with col5:
        categoria_extra = st.selectbox("Categoría", ["Interno", "Externo"])

    # Botón para añadir material
    if st.button("➕ Añadir material"):
        if descripcion_extra and cantidad_extra > 0 and precio_unit_extra > 0:
            nuevo_material = pd.DataFrame({
                "Referencia": [referencia_extra],
                "Descripcion": [descripcion_extra],
                "Cantidad": [cantidad_extra],
                "PrecioUnit": [precio_unit_extra],
                "Categoria": [categoria_extra],
                "PrecioTotal": [cantidad_extra * precio_unit_extra]
            })

            # Guardar en sesión temporal (para persistir hasta recargar)
            if "materiales_adicionales" not in st.session_state:
                st.session_state["materiales_adicionales"] = nuevo_material
            else:
                st.session_state["materiales_adicionales"] = pd.concat(
                    [st.session_state["materiales_adicionales"], nuevo_material],
                    ignore_index=True
                )
            st.success(f"✅ Material '{descripcion_extra}' añadido correctamente.")
        else:
            st.warning("⚠️ Debes introducir una descripción y valores válidos de cantidad y precio.")

# Recuperar materiales adicionales si existen
df_adicionales = st.session_state.get("materiales_adicionales", pd.DataFrame(
    columns=["Referencia","Descripcion", "Cantidad", "PrecioUnit", "Categoria", "PrecioTotal"]
))

# Mostrar materiales añadidos
if not df_adicionales.empty:
    st.markdown("### 🧾 Materiales adicionales añadidos")
    st.dataframe(
        df_adicionales[["Referencia","Descripcion", "Cantidad", "PrecioUnit", "PrecioTotal", "Categoria"]],
        use_container_width=True,
        hide_index=True
    )

# Incorporar los materiales adicionales al dataframe principal
if not df_adicionales.empty:
    df_filtrado = pd.concat([df_filtrado, df_adicionales], ignore_index=True)



# =====================
# 💰 CALCULAR PRECIOS
# =====================
df_filtrado["PrecioTotal"] = df_filtrado["Cantidad"] * df_filtrado["PrecioUnit"]

df_interno = df_filtrado[df_filtrado["Categoria"] == "Interno"]
df_externo = df_filtrado[df_filtrado["Categoria"] == "Externo"]


# =====================
# 📊 MOSTRAR RESULTADOS LISTADO DE MATERIALES
# =====================
st.markdown("## 1. Materiales")
st.markdown("### Material Interno")
st.dataframe(
    df_interno.reset_index(drop=True)[["Referencia","Descripcion", "Cantidad", "PrecioUnit", "PrecioTotal"]],
    use_container_width=True,
    hide_index=True
)
st.markdown(f"**Subtotal Interno:** {df_interno['PrecioTotal'].sum():,.2f} €")

st.markdown("### Material Externo")
st.dataframe(
    df_externo.reset_index(drop=True)[["Referencia","Descripcion", "Cantidad", "PrecioUnit", "PrecioTotal"]],
    use_container_width=True, hide_index=True
)
st.markdown(f"**Subtotal Externo:** {df_externo['PrecioTotal'].sum():,.2f} €")
st.markdown(f"**Total Material:** {df_filtrado['PrecioTotal'].sum():,.2f} €")


# =====================
# 👷 3. MANO DE OBRA
# =====================

st.markdown("---")
st.markdown("## 2. Mano de obra")

# --- Cálculo de días laborales según red y función ---

def calcular_dias_tecnico_electromecanico(red, funcion, num_maq):
    if red == "Profibus":
        if funcion == "Vision":
            return 1 * num_maq
        elif funcion == "Vision&Control":
            return 1 * num_maq + 2
    elif red == "Profinet":
        if funcion == "Vision":
            return 1 * num_maq
        elif funcion == "Vision&Control":
            return 0.5 * num_maq + 2
    return 0

def calcular_dias_tecnico_control(red, funcion, num_maq):
    # 🔹 Caso especial: si son 5 o más máquinas, Vision&Control y Profibus → 5 días fijos
    if red == "Profibus" and funcion == "Vision&Control" and num_maq >= 5:
        return 5

    if red == "Profibus":
        if funcion == "Vision":
            return 0.5 * num_maq + 1
        elif funcion == "Vision&Control":
            return 1 * num_maq + 1

    elif red == "Profinet":
        if funcion == "Vision":
            return 0.5 * num_maq + 1
        elif funcion == "Vision&Control" & num_maq!=5:
            return 0.5 * num_maq + 1

    return 0



# --- Aplicar lógica ---
dias_electromecanico = calcular_dias_tecnico_electromecanico(red, funcion, num_maq)
dias_control = calcular_dias_tecnico_control(red, funcion, num_maq)
dias_electromecanico_of = 0
dias_control_of = 1
num_electromecanico = 2
num_tec_control = 1
horas_act = 8

# =====================
# 🧮 DÍAS ADICIONALES MANUALES
# =====================
st.markdown("### ➕ Días adicionales (ajuste manual)")

col1, col2, col3, col4 = st.columns(4)
with col1:
    dias_extra_electro = st.number_input(
        "Días adicionales en obra (Electromecánico)",
        min_value=0.0, step=0.25, value=0.0
    )
with col2:
    dias_extra_ctrl = st.number_input(
        "Días adicionales en obra (Control)",
        min_value=0.0, step=0.25, value=0.0
    )
with col3:
    dias_extra_electro_of = st.number_input(
        "Días adicionales en oficina (Electromecánico)",
        min_value=0.0, step=0.25, value=0.0
    )
with col4:
    dias_extra_ctrl_of = st.number_input(
        "Días adicionales en oficina (Control)",
        min_value=0.0, step=0.25, value=0.0
    )

# --- Sumar a los días calculados automáticamente ---
dias_electromecanico += dias_extra_electro
dias_control += dias_extra_ctrl
dias_electromecanico_of += dias_extra_electro_of
dias_control_of += dias_extra_ctrl_of

# --- Crear DataFrames para mostrar ---
df_mano_obra = pd.DataFrame({
    "Técnico": ["Técnico Electromecánico", "Técnico de Control"],
    "Nº Personas": [num_electromecanico, num_tec_control],
    "Días en Obra": [dias_electromecanico, dias_control],
    "Dias Oficina" : [dias_electromecanico_of, dias_control_of]
})

# --- Mostrar resultados ---
st.markdown("### Técnico Electromecánico")
st.dataframe(df_mano_obra.reset_index(drop=True)[df_mano_obra["Técnico"] == "Técnico Electromecánico"][["Nº Personas", "Días en Obra","Dias Oficina"]],
             use_container_width=True,hide_index=True)

st.markdown("### Técnico de Control")
st.dataframe(df_mano_obra.reset_index(drop=True)[df_mano_obra["Técnico"] == "Técnico de Control"][["Nº Personas", "Días en Obra","Dias Oficina"]],
         use_container_width=True,hide_index=True)


# =====================
# 💰 CÁLCULO DE COSTES AGRUPADOS POR TIPO
# =====================

def obtener_tarifa(df, tecnico, palabra_clave):
    """Devuelve la tarifa y tipo correspondientes según el técnico y la descripción."""
    filtro = (
        (df["Técnico"].isin([tecnico, "Ambos"])) &
        (df["DESCRIPCION"].str.contains(palabra_clave, case=False, na=False))
    )
    try:
        fila = df.loc[filtro].iloc[0]
        return fila["TARIFA"], fila["Tipo"]
    except IndexError:
        return 0, None


# --- Obtener tarifas y tipos ---
tarifa_hora_electro, tipo_hora_electro = obtener_tarifa(df_tarifas, "Técnico Electromecánico", "HORA TRABAJO NORMAL")
tarifa_hora_ctrl, tipo_hora_ctrl = obtener_tarifa(df_tarifas, "Técnico de Control", "HORA TRABAJO NORMAL")
tarifa_hora_despl_electro, tipo_hora_despl_electro = obtener_tarifa(df_tarifas, "Técnico Electromecánico", "DESPLAZAMIENTO TECNICO")
tarifa_hora_despl_ctrl, tipo_hora_despl_ctrl = obtener_tarifa(df_tarifas, "Técnico de Control", "DESPLAZAMIENTO INGENIERO")

tarifa_media_dieta, tipo_media_dieta = obtener_tarifa(df_tarifas, "Ambos", "MEDIA DIETA")
tarifa_dieta_completa, tipo_dieta_completa = obtener_tarifa(df_tarifas, "Ambos", "DIETA COMPLETA")
tarifa_pernocta, tipo_pernocta = obtener_tarifa(df_tarifas, "Ambos", "PERNOCTACION")
tarifa_km, tipo_km = obtener_tarifa(df_tarifas, "Ambos", "KILOMETRO")
tarifa_frio, tipo_frio = obtener_tarifa(df_tarifas, "Ambos", "FRIO")
# --- NUEVAS TARIFAS PARA HORAS FESTIVAS Y NOCTURNAS ---
tarifa_plus_nocturnidad, tipo_plus_nocturnidad = obtener_tarifa(df_tarifas, "Ambos", "PLUS POR JORNADA DE NOCTURNIDAD")
tarifa_nocturna_electro, tipo_nocturna_electro = obtener_tarifa(df_tarifas, "Técnico Electromecánico", "HORA TRABAJO NOCTURNA/FESTIVA TECNICO")
tarifa_nocturna_ctrl, tipo_nocturna_ctrl = obtener_tarifa(df_tarifas, "Técnico de Control", "HORA TRABAJO NOCTURNA/FESTIVA INGENIERO")
tarifa_plus_turnicidad, tipo_plus_turnicidad = obtener_tarifa(df_tarifas, "Ambos", "PLUS POR JORNADA DE TURNICIDAD")
# --- NUEVAS TARIFAS PARA DESPLAZAMIENTOS ---
tarifa_avion, tipo_avion = obtener_tarifa(df_tarifas, "Ambos", "PRECIO AVION ESPAÑA")
tarifa_alquiler, tipo_alquiler = obtener_tarifa(df_tarifas, "Ambos", "PRECIO VEHICULO DE ALQUILER POR DIA")


# --- Cálculo de cantidades ---
cant_horas_electro = (dias_electromecanico+dias_electromecanico_of) * num_electromecanico * horas_act
cant_despl_electro = num_electromecanico * h_electro
cant_media_dieta_electro = dias_electromecanico * num_electromecanico if estancia_electro == "VIAJA" else 0
cant_dieta_completa_electro = dias_electromecanico * num_electromecanico if estancia_electro == "PERNOCTA" else 0
cant_pernocta_electro = dias_electromecanico * num_electromecanico if estancia_electro == "PERNOCTA" else 0
cant_km_electro = km_electro
cant_frio_electro = dias_electromecanico * num_electromecanico if frio_electro == "SI" else 0

cant_horas_ctrl = (dias_control+dias_control_of) * num_tec_control * horas_act
cant_despl_ctrl = num_tec_control * h_ctrl
cant_media_dieta_ctrl = dias_control * num_tec_control if estancia_ctrl == "VIAJA" else 0
cant_dieta_completa_ctrl = dias_control * num_tec_control if estancia_ctrl == "PERNOCTA" else 0
cant_pernocta_ctrl = dias_control * num_tec_control if estancia_ctrl == "PERNOCTA" else 0
cant_km_ctrl = km_ctrl
cant_frio_ctrl = dias_control * num_tec_control if frio_ctrl == "SI" else 0


# --- CÁLCULOS DE VIAJE Y VEHÍCULO ---
# Avión: se multiplica x2 por técnico si el viaje es en avión
coste_avion_electro = 0
coste_avion_ctrl = 0
if tipo_viaje_electro == "AVION":
    coste_avion_electro = tarifa_avion * num_electromecanico * 2
if tipo_viaje_ctrl == "AVION":
    coste_avion_ctrl = tarifa_avion * num_tec_control * 2

# Vehículo de alquiler: coste por días
coste_alquiler_electro = dias_alquiler_electro * tarifa_alquiler
coste_alquiler_ctrl = dias_alquiler_ctrl * tarifa_alquiler



# --- Definición de estructura de costes con tipo ---
datos_coste = [
    # Tipo Horas Laborales
    {"Tipo": tipo_hora_electro, "Concepto": "Hora trabajo normal", "Electromecánico": cant_horas_electro * tarifa_hora_electro, "Control": cant_horas_ctrl * tarifa_hora_ctrl},
    {"Tipo": tipo_hora_despl_electro, "Concepto": "Hora desplazamiento", "Electromecánico": cant_despl_electro * tarifa_hora_despl_electro, "Control": cant_despl_ctrl * tarifa_hora_despl_ctrl},
    # Tipo Complementos
    {"Tipo": tipo_media_dieta, "Concepto": "Media dieta", "Electromecánico": cant_media_dieta_electro * tarifa_media_dieta, "Control": cant_media_dieta_ctrl * tarifa_media_dieta},
    {"Tipo": tipo_dieta_completa, "Concepto": "Dieta completa", "Electromecánico": cant_dieta_completa_electro * tarifa_dieta_completa, "Control": cant_dieta_completa_ctrl * tarifa_dieta_completa},
    {"Tipo": tipo_pernocta, "Concepto": "Pernocta", "Electromecánico": cant_pernocta_electro * tarifa_pernocta, "Control": cant_pernocta_ctrl * tarifa_pernocta},
    {"Tipo": tipo_km, "Concepto": "Kilómetros", "Electromecánico": cant_km_electro * tarifa_km, "Control": cant_km_ctrl * tarifa_km},
    {"Tipo": tipo_frio, "Concepto": "Trabajo en frío", "Electromecánico": cant_frio_electro * tarifa_frio, "Control": cant_frio_ctrl * tarifa_frio},
    # --- NUEVOS CONCEPTOS: Desplazamientos (avión y alquiler vehículo) ---
    {"Tipo": "Complementos", "Concepto": "Viaje en avión (ida y vuelta)",
    "Electromecánico": coste_avion_electro,
    "Control": coste_avion_ctrl},

    {"Tipo": "Complementos", "Concepto": "Vehículo de alquiler (por días)",
    "Electromecánico": coste_alquiler_electro,
    "Control": coste_alquiler_ctrl},
    # --- NUEVOS CONCEPTOS: Horas especiales (Festivo y Nocturno) ---
    # 1️⃣ Trabajo en festivo → se calcula por horas
    {"Tipo": "Complementos", "Concepto": "Trabajo en festivo",
    "Electromecánico": horas_festivo_electro * tarifa_nocturna_electro*num_electromecanico,
    "Control": horas_festivo_ctrl * tarifa_nocturna_ctrl*num_tec_control},

    # 2️⃣ Trabajo en nocturno → se calcula por horas + PLUS por jornada (nocturnidad + turnicidad)
    {
    "Tipo": "Complementos",
    "Concepto": "Trabajo en nocturno",
    "Electromecánico": (
      horas_nocturno_electro * tarifa_nocturna_electro*num_electromecanico
      + (
         (horas_nocturno_electro / horas_act)  # convertir horas en días
         * num_electromecanico
         * (tarifa_plus_nocturnidad + tarifa_plus_turnicidad)
        )
     ),
    "Control": (
      horas_nocturno_ctrl * tarifa_nocturna_ctrl* num_tec_control
      + (
         (horas_nocturno_ctrl / horas_act)
         * num_tec_control
         * (tarifa_plus_nocturnidad + tarifa_plus_turnicidad)
        )
     ),
   },

          ]

df_costes = pd.DataFrame(datos_coste)

# --- Subtotales por Tipo ---
df_subtotales = df_costes.groupby("Tipo", dropna=False)[["Electromecánico", "Control"]].sum().reset_index()
df_subtotales["Concepto"] = "Subtotal " + df_subtotales["Tipo"]
df_subtotales = df_subtotales[["Tipo", "Concepto", "Electromecánico", "Control"]]

# --- Total general ---
df_total = pd.DataFrame([{
    "Tipo": "",
    "Concepto": "TOTAL GENERAL",
    "Electromecánico": df_costes["Electromecánico"].sum(),
    "Control": df_costes["Control"].sum()
}])

# --- Unión final ---
df_final = pd.concat([df_costes, df_subtotales, df_total], ignore_index=True)

# --- Mostrar resultados ---
st.markdown("### Costes Mano de obra")
for col in ["Electromecánico", "Control"]:
    df_final[col] = pd.to_numeric(df_final[col], errors="coerce")
df_final[["Electromecánico", "Control"]] = df_final[["Electromecánico", "Control"]].fillna(0.0)
styler = df_final.style.format({
    "Electromecánico": "{:,.2f} €",
    "Control": "{:,.2f} €"
})
st.dataframe(styler, use_container_width=True,hide_index=True)

# --- Totales por Tipo (Horas Laborales / Complementos) ---
totales_por_tipo = (
    df_subtotales.groupby("Tipo")[["Electromecánico", "Control"]]
    .sum()
    .sum(axis=1)
    .to_dict()
)

total_horas_laborales = totales_por_tipo.get("Horas Laborales", 0)
total_complementos = totales_por_tipo.get("Complementos", 0)

# --- Mostrar resumen en markdown ---
st.markdown(f"**Total Horas Laborales:** {total_horas_laborales:,.2f} €")
st.markdown(f"**Total Complementos:** {total_complementos:,.2f} €")
st.markdown(f"**Total Mano de Obra:** {total_horas_laborales + total_complementos:,.2f} €")


# =====================
# 🧠 4. HORAS DE INGENIERÍA
# =====================

# Definir tarifa e horas según tipo de instalación
tarifa_ingenieria = 190

if funcion == "Vision":
    horas_ingenieria = 40
elif funcion == "Vision&Control":
    horas_ingenieria = 64
else:
    horas_ingenieria = 0

# Calcular coste total
coste_ingenieria = horas_ingenieria * tarifa_ingenieria

# Mostrar resultados
st.markdown("---")
st.markdown("## 3. Horas de ingeniería")

df_ingenieria = pd.DataFrame({
    "Concepto": ["Horas de ingeniería"],
    "Horas": [horas_ingenieria],
    "Tarifa (€/h)": [tarifa_ingenieria],
    "Coste Total (€)": [coste_ingenieria]
})

st.dataframe(
    df_ingenieria.style.format({
        "Tarifa (€/h)": "{:,.2f} €",
        "Coste Total (€)": "{:,.2f} €"
    }),
    use_container_width=True,hide_index=True
)

# =====================
# 📊 4. RESUMEN DE OFERTA (con subdivisión de mano de obra)
# =====================

st.markdown("---")
st.header("📊 4. Resumen de oferta")

# --- Datos generales ---
st.subheader("Datos generales")
col1, col2 = st.columns(2)
with col1:
    fecha_decision = st.date_input("📅 Fecha de decisión")
    poblacion = st.text_input("🏙️ Población", "")
with col2:
    cliente = st.text_input("👤 Cliente", "")
    delegacion = st.text_input("🏢 Delegación", "")

# --- Sección de descuentos ---
st.subheader("Descuentos aplicables (%)")

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    descuento_interno = st.number_input("Desc. Materiales Internos (%)", min_value=0.0, max_value=100.0, value=10.0, step=0.5)
with col2:
    descuento_externo = st.number_input("Desc. Materiales Externos (%)", min_value=0.0, max_value=100.0, value=55.0, step=0.5)
with col3:
    descuento_horas = st.number_input("Desc. Mano de Obra - Horas Laborales (%)", min_value=0.0, max_value=100.0, value=30.0, step=0.5)
with col4:
    descuento_complementos = st.number_input("Desc. Mano de Obra - Complementos (%)", min_value=0.0, max_value=100.0, value=30.0, step=0.5)
with col5:
    descuento_ingenieria = st.number_input("Desc. Horas de Ingeniería (%)", min_value=0.0, max_value=100.0, value=30.0, step=0.5)

# --- Calcular totales por grupo ---
total_interno = df_interno["PrecioTotal"].sum()
total_externo = df_externo["PrecioTotal"].sum()

# Calcular subtotales mano de obra por tipo
#subtotal_horas = df_final.loc[df_final["Tipo"] == "Horas Laborales", ["Electromecánico", "Control"]].sum().sum()
#subtotal_complementos = df_final.loc[df_final["Tipo"] == "Complementos", ["Electromecánico", "Control"]].sum().sum()
#total_mano_obra = subtotal_horas + subtotal_complementos

# ✅ Calcular subtotales de mano de obra por tipo directamente desde df_costes
subtotal_horas = (
    df_costes[df_costes["Tipo"] == "Horas Laborales"][["Electromecánico", "Control"]]
    .sum()
    .sum()
)

subtotal_complementos = (
    df_costes[df_costes["Tipo"] == "Complementos"][["Electromecánico", "Control"]]
    .sum()
    .sum()
)

# Asegurar valores numéricos (evita NaN)
subtotal_horas = float(subtotal_horas) if pd.notna(subtotal_horas) else 0.0
subtotal_complementos = float(subtotal_complementos) if pd.notna(subtotal_complementos) else 0.0

# Total mano de obra = suma de ambos
total_mano_obra = subtotal_horas + subtotal_complementos



# Totales de ingeniería
total_ingenieria = coste_ingenieria

# --- Aplicar descuentos ---
neto_interno = total_interno * (1 - descuento_interno / 100)
neto_externo = total_externo * (1 - descuento_externo / 100)
neto_horas = subtotal_horas * (1 - descuento_horas / 100)
neto_complementos = subtotal_complementos * (1 - descuento_complementos / 100)
neto_ingenieria = total_ingenieria * (1 - descuento_ingenieria / 100)

# Totales generales
total_tarifa = total_interno + total_externo + subtotal_horas + subtotal_complementos + total_ingenieria
total_neto = neto_interno + neto_externo + neto_horas + neto_complementos + neto_ingenieria
total_descuento = ((total_tarifa - total_neto) / total_tarifa) * 100

# --- Crear tabla resumen ---
df_resumen = pd.DataFrame({
    "Concepto": [
        "1. Materiales Internos",
        "2. Materiales Externos",
        "3.1 Mano de Obra - Horas Laborales",
        "3.2 Mano de Obra - Complementos",
        "4. Ingeniería"
    ],
    "Tarifa (€)": [
        total_interno,
        total_externo,
        subtotal_horas,
        subtotal_complementos,
        total_ingenieria
    ],
    "Descuento (%)": [
        descuento_interno,
        descuento_externo,
        descuento_horas,
        descuento_complementos,
        descuento_ingenieria
    ],
    "Neto Cliente (€)": [
        neto_interno,
        neto_externo,
        neto_horas,
        neto_complementos,
        neto_ingenieria
    ]
})

# --- Mostrar resumen ---
st.markdown("### 🧾 Resumen de importes")
st.dataframe(
    df_resumen.style.format({
        "Tarifa (€)": "{:,.2f} €",
        "Descuento (%)": "{:.2f} %",
        "Neto Cliente (€)": "{:,.2f} €"
    }).hide(axis="index"),
    use_container_width=True,hide_index=True
)

# --- Totales ---
st.markdown("### 💵 Totales generales")
col1, col2, col3 = st.columns([1, 1, 1])
col1.metric("Total Tarifa", f"{total_tarifa:,.2f} €")
col2.metric("Descuento medio", f"{total_descuento:,.2f} %")
col3.metric("Total Neto Cliente", f"{total_neto:,.2f} €")


# =====================
# 💾 EXPORTAR A EXCEL (Opción 2 con unión de filas 3.1 y 3.2)
# =====================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

output = BytesIO()
wb = Workbook()
ws = wb.active
ws.title = "Resumen Oferta"

# --- Estilos ---
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# --- Título ---
titulo_texto = f"OFERTA: Cámaras para visualizar y controlar desde {delegacion or 'CPC'} ({poblacion or ''})"
ws["A1"] = titulo_texto
ws.merge_cells("A1:E1")
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = center

# --- Encabezados ---
headers = ["Concepto", "Detalle", "Tarifa (€)", "Descuento (%)", "Neto Cliente (€)"]
for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=6, column=col, value=text)
    cell.font = bold
    cell.alignment = center
    cell.fill = gray_fill
    cell.border = border

# --- Construcción de detalles dinámicos ---
def generar_detalle(concepto):
    if concepto == "1. Materiales Internos":
        if not df_interno.empty:
            return "\n".join(
                f"- {row['Descripcion']} ({row['Cantidad']})"
                for _, row in df_interno.iterrows()
            )
        else:
            return "Sin materiales internos"

    elif concepto == "2. Materiales Externos":
        if not df_externo.empty:
            return "\n".join(
                f"- {row['Descripcion']} ({row['Cantidad']})"
                for _, row in df_externo.iterrows()
            )
        else:
            return "Sin materiales externos"

    elif concepto in ["3.1 Mano de Obra - Horas Laborales", "3.2 Mano de Obra - Complementos"]:
        return (
            f"Técnico Electromecánico: {num_electromecanico} pers. x {dias_electromecanico} días\n"
            f"Técnico de Control: {num_tec_control} pers. x {dias_control} días"
        )

    elif concepto == "4. Ingeniería":
        return f"Horas de ingeniería: {horas_ingenieria}"

    else:
        return ""

# --- Datos dinámicos ---
start_row = 7
row_map = {}  # para guardar índices y luego fusionar
for i, (_, fila) in enumerate(df_resumen.iterrows(), start=start_row):
    concepto = fila["Concepto"]
    detalle = generar_detalle(concepto)

    ws[f"A{i}"] = concepto

    # 💡 Caso especial: unificar detalle para 3.1 y 3.2
    if concepto == "3.1 Mano de Obra - Horas Laborales":
        ws[f"B{i}"] = detalle
        row_map["manodeobra_inicio"] = i  # guardamos para fusionar después
    elif concepto == "3.2 Mano de Obra - Complementos":
        ws[f"B{i}"] = ""  # esta queda vacía
        row_map["manodeobra_fin"] = i
    else:
        ws[f"B{i}"] = detalle

    ws[f"C{i}"] = float(fila["Tarifa (€)"])
    ws[f"D{i}"] = float(fila["Descuento (%)"]) / 100.0  # ✅ formato decimal %
    ws[f"E{i}"] = float(fila["Neto Cliente (€)"])

    # Aplicar formato
    for col in "ABCDE":
        c = ws[f"{col}{i}"]
        c.border = border
        c.alignment = left if col in ["A", "B"] else right
        if col in ["C", "E"]:
            c.number_format = "#,##0.00 €"
        elif col == "D":
            c.number_format = "0.00%"

# --- Fusionar celdas de Detalle (Mano de Obra) ---
if "manodeobra_inicio" in row_map and "manodeobra_fin" in row_map:
    ws.merge_cells(
        start_row=row_map["manodeobra_inicio"],
        start_column=2,
        end_row=row_map["manodeobra_fin"],
        end_column=2
    )

# --- Total general ---
total_row = start_row + len(df_resumen)
ws[f"A{total_row}"] = "TOTAL GENERAL"
ws[f"C{total_row}"] = total_tarifa
ws[f"D{total_row}"] = total_descuento / 100.0
ws[f"E{total_row}"] = total_neto

for col in "ABCDE":
    c = ws[f"{col}{total_row}"]
    c.font = bold
    c.fill = yellow_fill
    c.border = border
    c.alignment = right
    if col in ["C", "E"]:
        c.number_format = "#,##0.00 €"
    elif col == "D":
        c.number_format = "0.00%"

# --- Ajuste de anchos ---
col_widths = {"A": 35, "B": 55, "C": 18, "D": 18, "E": 20}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Ajustar altura automática de fila para texto multilinea
for row in ws.iter_rows(min_row=start_row, max_row=total_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="center")

# --- Guardar ---
wb.save(output)
output.seek(0)

# --- Descargar ---
st.download_button(
    label="💾 Descargar oferta completa (Excel)",
    data=output,
    file_name=f"Oferta_Resumen_{delegacion or 'General'}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
